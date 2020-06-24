import openpyxl, statistics

wb = openpyxl.load_workbook('C:/Users/jatin/Downloads/Book1.xlsx')
print(wb)
sheet = wb.get_sheet_by_name("Russia")
ws = wb.active
a = sheet["A" + "2"].value
print(a[5:7])
print(type(a))
d1 = 0
d2 = 0


def check_record_by_date(d):
    for row2 in ws.rows:
        if row2[0].value == d:
            if row2[1].value is not None:
                return True
    return False


def get_avg_temp_by_date(d):
    for row1 in ws.rows:
        if row1[0].value == d:
            if row[1].value is not None:
                return row1[1].value
    return None


def get_avg_temp_uncertain_by_date(d):
    for row1 in ws.rows:
        if row1[0].value == d:
            return row1[2].value
    return None


def travel_date(d, backward=True, by=1):
    year = d[0:4]
    if backward:
        g_year = int(year) - by
    if not backward:
        g_year = int(year) + by
    a_date = d.replace(year, str(g_year))
    return a_date


states = ["Chukot", "Chuvash", "City Of St. Petersburg", "Dagestan", "Evenk", "Gorno Altay", "Ingush"]
row_no = 0
for state in states:
    for row in ws.rows:
        row_no += 1
        if row[3].value == state and row[4].value == "Russia":
            if row[2].value is None:
                cleaned_value = "0"
                mis_date = str(row[0].value)
                print(mis_date)
                new_date1 = travel_date(mis_date)
                new_date2 = travel_date(mis_date, by=2)
                print(new_date1)
                print(new_date2)
                record_exist1 = check_record_by_date(new_date1)
                print(record_exist1)
                record_exist2 = check_record_by_date(new_date2)
                print(record_exist2)
                if record_exist1 is False and record_exist2 is False:
                    print(state)
                    print("inside double false")
                    new_date1 = travel_date(mis_date, backward=False)
                    print(new_date1)
                    new_date2 = travel_date(mis_date, backward=False, by=2)
                    print(new_date2)
                    record_exist1 = check_record_by_date(new_date1)
                    print(record_exist1)
                    record_exist2 = check_record_by_date(new_date2)
                    print(record_exist2)
                if record_exist1 is False and record_exist2 is True:
                    print(state)
                    print("Here First record is false and second True")
                    cleaned_value = get_avg_temp_by_date(new_date2)
                    #uncertain_cleaned_value = get_avg_temp_uncertain_by_date(new_date2)
                    print(cleaned_value)
                    #print(uncertain_cleaned_value)
                if record_exist1 is True and record_exist2 is False:
                    print(state)
                    print("Here first record is True and second is False")
                    cleaned_value = get_avg_temp_by_date(new_date1)
                    #uncertain_cleaned_value = get_avg_temp_uncertain_by_date(new_date1)
                    print(cleaned_value)
                    #print(uncertain_cleaned_value)
                if record_exist1 is True and record_exist2 is True:
                    print(state)
                    print("Here in double True")
                    data1 = get_avg_temp_by_date(new_date1)
                    #u_data1 = get_avg_temp_uncertain_by_date(new_date1)
                    print("Data1" + "///" + str(data1) + "///" + "///" + str(new_date1))
                    data2 = get_avg_temp_by_date(new_date2)
                    #u_data2 = get_avg_temp_uncertain_by_date(new_date2)
                    print("Data2" + "///" + str(data2) + "///" + "///" + str(new_date2))
                    cleaned_value = "check"
                    #uncertain_cleaned_value = "check"
                    if data1 is not None and data2 is not None:
                        cleaned_value = statistics.mean([data1, data2])
                    #if u_data1 is not None and u_data2 is not None:
                    #   uncertain_cleaned_value = statistics.mean([u_data1, u_data2])

                sheet.cell(row=row_no, column=2).value = cleaned_value
                #sheet.cell(row=row_no, column=3).value = uncertain_cleaned_value
wb.save("C:/Users/jatin/Downloads/updated_Russia4.xlsx")

#        year = mis_date[2:4]
#        prev_year = int(year) - 1
#        f_date = mis_date.replace(year, str(prev_year))
#        result = check_record_by_date(f_date)
#        if result == False:
# Check ahead month
#            next_year
#        for row1 in ws.rows:
#            found = 0
#            if row1[0].value == f_date:
#                print("yes")
#                found = 1
#            if found == 0:
#                prev_year
# for row in range(2, sheet.max_row + 1):
#    value = sheet["B" + str(row)].value
#    if value != None:
#        pass
#    elif value == None:
#        pass

# print(sheet['B' + "3"].value)
